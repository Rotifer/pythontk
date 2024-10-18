import os
import duckdb
import openpyxl as pyxl
import openpyxl.utils  as pyxl_utils #import get_column_letter
import pandas as pd

class ExcelInfo:
    def __init__(self, excel_filepath):
        if os.path.exists(excel_filepath):
            self.excel_filepath = excel_filepath
        else:
            raise IOError(f"Given Excel file {excel_filepath} does not exist!")
        self.workbook = pyxl.load_workbook(filename=self.excel_filepath)
        self._sheet_names = []
        self.wb = pyxl.load_workbook(excel_filepath)
        self._wb_properties = self.extract_workbook_properties()
        self._wb_properties["filepath"] = self.excel_filepath
    
    def extract_workbook_properties(self):
        """Return all workbook properties that have been set as a dictionary.

        Date property values are converted to 'YYYY-MM-DD' format.
        """
        property_names = ["creator", "title", "description", "subject", "identifier",
                          "language", "created","modified", "lastModifiedBy", "category", 
                          "contentStatus", "version", "revision", "keywords", "lastPrinted"]
        wb_properties = {}
        for property_name in property_names:
            property_value = getattr(self.wb.properties, property_name)
            if property_value:
                if property_name in ["created", "modified", "lastPrinted"]:
                    property_value = property_value.strftime("%Y-%m-%d")
                wb_properties[property_name] = property_value
        return wb_properties

    @property
    def sheet_names(self):
        """
        """
        self._sheet_names = self.workbook.sheetnames
        return self._sheet_names
    
    def _get_sheet_object(self, sheet_name):
        if sheet_name not in self.sheet_names:
            raise ValueError(f"{sheet_name} does not exist!")
        sheet = self.workbook[sheet_name]
        return sheet
    
    def get_row_values(self, sheet_name, row_number=1):
        """_summary_

        Args:
            row_number (int, optional): _description_. Defaults to 1.
        """
        sheet = self._get_sheet_object(sheet_name)
        row_values = [cell.value for cell in sheet[row_number]]
        return row_values
    
    def get_last_row_for_sheet(self, sheet_name):
        sheet = self._get_sheet_object(sheet_name)
        return sheet.max_row
    
    def get_last_column_for_sheet(self, sheet_name):
        sheet = self._get_sheet_object(sheet_name)
        return sheet.max_column
    
    def get_sheet_usedrange_address(self, sheet_name):
        max_column = self.get_last_column_for_sheet(sheet_name)
        max_row = self.get_last_row_for_sheet(sheet_name)
        last_column_letter = pyxl_utils.get_column_letter(max_column)
        usedrange_address = f"A1:{last_column_letter + str(max_row)}"
        return usedrange_address
    
    def get_sheet_column_types(self, sheet_name, column_names_row=1, nrows=10):
        """Determine the DuckDB data types of the input Excel sheet columns.

        By default only the first 10 rows are tested but this can be increased.
        Remove the index column created by Pandas before returning the list.

        Args:
            sheet_name (_type_): _description_
            column_names_row (int, optional): _description_. Defaults to 1.
            nrows (int, optional): _description_. Defaults to 10.

        Returns:
            list(tuple): Each tuple is composed of the column name and column type.
        """
        df = pd.read_excel(self.excel_filepath,
                           sheet_name=sheet_name,
                           skiprows=column_names_row - 1,
                           header=column_names_row - 1,
                           nrows=nrows)
        conn = duckdb.connect(database=':memory:', read_only=False)
        df.to_sql("temp_tbl", conn)
        sql = "DESCRIBE temp_tbl"
        result = conn.execute(sql)
        names_types = [(row[0], row[1]) for row in result.fetchall()]
        conn.close()
        names_types.pop(0)
        return names_types

    
if __name__ == "__main__":
    excel_filepath = "TheraSAbDab_SeqStruc_OnlineDownload.xlsx"
    sheet_name = "TheraSAbDab_June24"
    excel_info = ExcelInfo(excel_filepath)
    #print(xl_info.sheet_names)
    #print(xl_info.get_row_values(sheet_name, 1))
    #print(xl_info.get_sheet_usedrange_address(sheet_name))
    print(excel_info.extract_workbook_properties())
    print(excel_info.get_sheet_column_types(sheet_name))